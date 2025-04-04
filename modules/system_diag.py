from fabric import Connection
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

class SystemDiagSSH:
    def __init__(self, host, user, password):
        self.host = host
        self.user = user
        self.password = password
        self.conn = None
        self.results = {}
        self.failures = []

    def connect(self):
        try:
            self.conn = Connection(
                host=self.host,
                user=self.user,
                connect_kwargs={"password": self.password},
                connect_timeout=10
            )
            self.conn.open()
            return True
        except Exception as e:
            print(f"\n❌ Connexion SSH échouée : {e}")
            return False

    def run_command(self, label, command, parse=False):
        try:
            result = self.conn.run(command, hide=True, warn=True, encoding='utf-8')
            output = result.stdout.strip()

            if parse:
                if label == "CPU":
                    parsed = {}
                    for line in output.splitlines():
                        if ":" in line:
                            key, value = line.split(":", 1)
                            parsed[key.strip()] = value.strip()
                    self.results[label] = parsed

                elif label == "RAM":
                    parts = output.split()
                    if len(parts) >= 2:
                        self.results[label] = {"total": parts[1]}
                    else:
                        self.results[label] = {"total": "N/A"}

                elif label == "Bigs process":
                    lines = output.splitlines()[1:]
                    processes = []
                    for line in lines:
                        parts = line.split(None, 2)
                        if len(parts) == 3:
                            processes.append({"PID": parts[0], "COMMAND": parts[1], "%MEM": parts[2]})
                    self.results[label] = processes

                elif label == "Env variable":
                    env = {}
                    for line in output.splitlines():
                        if "=" in line:
                            k, v = line.split("=", 1)
                            env[k.strip()] = v.strip()
                    self.results[label] = env

                elif label == "Disk usage":
                    lines = output.splitlines()[1:]
                    disks = {}
                    for line in lines:
                        parts = line.split()
                        if len(parts) == 6:
                            disks[parts[5]] = {
                                "size": parts[1], "used": parts[2],
                                "avail": parts[3], "pcent": parts[4]
                            }
                    self.results[label] = disks

                elif label == "Disks":
                    lines = output.splitlines()[1:]
                    disks = []
                    for line in lines:
                        parts = line.split()
                        if len(parts) >= 4:
                            name_clean = parts[0].replace("├─", "").replace("└─", "").replace("│", "").strip("─ ")
                            disks.append(f"{name_clean}, SIZE: {parts[1]}, TYPE: {parts[2]}, MOUNTPOINT: {parts[3] if len(parts) > 3 else ''}")
                    self.results[label] = disks

                elif label == "Network interfaces":
                    interfaces = output.splitlines()
                    self.results[label] = interfaces

                elif label == "Boot time":
                    parts = output.split()
                    self.results[label] = {"Boot time": " ".join(parts) if parts else "N/A"}

                else:
                    self.results[label] = output
            else:
                self.results[label] = output

        except Exception:
            self.results[label] = None
            self.failures.append(label)

    def collect_info(self):
        cmds = {
            "OS": ("uname -srmo", False),
            "CPU": ("lscpu | grep -E \"Nom|Architecture|Processeur\"", True),
            "RAM": ("free -h | grep Mem", True),
            "Bigs process": ("ps -eo pid,comm,%mem --sort=-%mem | head -n 6", True),
            "Env variable": ("printenv | grep -E \"^(PATH|HOME|USER)=\"", True),
            "Disks": ("lsblk -o NAME,SIZE,TYPE,MOUNTPOINT", True),
            "Disk usage": ("df -h --output=source,size,used,avail,pcent,target", True),
            "Network interfaces": ("ip -o link show | awk -F': ' '{print $2}'", True),
            "Boot time": ("who -b | awk '{print $3, $4}'", True)
        }
        for label, (cmd, parse) in cmds.items():
            self.run_command(label, cmd, parse=parse)

    def print_summary(self):
        if self.failures:
            print("\n🚨 Certaines informations n'ont pas pu être récupérées :")
            for label in self.failures:
                print(f"  ❌ {label}")
        else:
            print("\n✅ Diagnostic complété avec succès !")

    def export_to_excel(self, path):
        data = {label: self.results.get(label, "N/A") for label in [
            "OS", "CPU", "RAM", "Bigs process", "Env variable",
            "Disks", "Disk usage", "Network interfaces", "Boot time"
        ]}
        for key, val in data.items():
            if isinstance(val, dict):
                if key == "Disk usage":
                    data[key] = '\n'.join(
                        f"{mount} > " + ', '.join(f"{k} : {v}" for k, v in details.items())
                        for mount, details in val.items()
                    )
                else:
                    data[key] = '\n'.join(f"{k} : {v}" for k, v in val.items())
            elif isinstance(val, list):
                data[key] = '\n'.join(
                    item if isinstance(item, str)
                    else ', '.join(f"{k}: {v}" for k, v in item.items())
                    for item in val
                )

        data = {"IP target": self.host, **data}
        df = pd.DataFrame([data])

        sheet_name = "System status"
        file_exists = os.path.exists(path)
        if file_exists:
            with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(path, engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Ajuster l'alignement dans les cellules (hors entête)
        wb = load_workbook(path)
        ws = wb[sheet_name]
        alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = alignment
        wb.save(path)
